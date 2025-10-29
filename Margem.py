import tkinter as tk
from tkinter import messagebox, filedialog
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

# Função para calcular lucro e margem
def calcular_lucro():
    try:
        preco_compra = float(entry_compra.get())
        preco_venda = float(entry_venda.get())
        custos_adicionais = float(entry_custos.get() or 0)
        custo_frete = float(entry_frete.get() or 0)
        
        lucro = preco_venda - preco_compra - custos_adicionais - custo_frete
        margem_lucro = (lucro / preco_venda) * 100 if preco_venda > 0 else 0
        
        label_resultado.config(text=f"Lucro Líquido: R$ {lucro:.2f}\nMargem de Lucro: {margem_lucro:.2f}%")
        
        # Armazenar valores para salvamento
        global valores_calculados
        valores_calculados = [preco_compra, preco_venda, custos_adicionais, custo_frete, lucro, margem_lucro]
    except ValueError:
        messagebox.showerror("Erro", "Digite valores numéricos válidos!")

# Função para salvar dados na planilha com gráfico
def salvar_dados():
    nome_produto = entry_nome.get()
    if not nome_produto or not valores_calculados:
        messagebox.showerror("Erro", "Calcule os valores e insira o nome do produto!")
        return
    
    try:
        # Carregar ou criar planilha
        try:
            wb = openpyxl.load_workbook('planilha.xlsx')
            sheet = wb.active
        except:
            wb = Workbook()
            sheet = wb.active
            sheet.title = 'Resumo da Calculadora de Lucro'
            # Cabeçalhos
            sheet['A1'] = 'Nome do Produto'
            sheet['B1'] = 'Preço de Compra'
            sheet['C1'] = 'Preço de Venda'
            sheet['D1'] = 'Custos Adicionais'
            sheet['E1'] = 'Custo do Frete'
            sheet['F1'] = 'Lucro Líquido'
            sheet['G1'] = 'Margem de Lucro (%)'
        
        # Adicionar nova linha
        row = sheet.max_row + 1
        sheet[f'A{row}'] = nome_produto
        sheet[f'B{row}'] = valores_calculados[0]
        sheet[f'C{row}'] = valores_calculados[1]
        sheet[f'D{row}'] = valores_calculados[2]
        sheet[f'E{row}'] = valores_calculados[3]
        sheet[f'F{row}'] = valores_calculados[4]
        sheet[f'G{row}'] = valores_calculados[5]
        
        # Gerar gráfico de barras para lucros (coluna F)
        chart = BarChart()
        chart.title = "Lucros Líquidos dos Produtos"
        chart.y_axis.title = "Lucro (R$)"
        chart.x_axis.title = "Produtos"
        
        # Referências: nomes (A2:A...) e lucros (F2:F...)
        data = Reference(sheet, min_col=6, min_row=2, max_row=row)  # Coluna F (lucros)
        cats = Reference(sheet, min_col=1, min_row=2, max_row=row)  # Coluna A (nomes)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)
        
        # Adicionar gráfico à planilha (posição arbitrária, ex.: I2)
        sheet.add_chart(chart, "I2")
        
        wb.save('planilha.xlsx')
        messagebox.showinfo("Sucesso", "Dados salvos e gráfico gerado em 'planilha.xlsx'!")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao salvar: {str(e)}")

# Função para carregar e exibir dados
def carregar_dados():
    try:
        dados = obter_dados_excel('planilha.xlsx')
        text_area.delete(1.0, tk.END)
        if dados:
            for linha in dados:
                text_area.insert(tk.END, f"Produto: {linha[0]}, Compra: R$ {linha[1]:.2f}, Venda: R$ {linha[2]:.2f}, Custos: R$ {linha[3]:.2f}, Frete: R$ {linha[4]:.2f}, Lucro: R$ {linha[5]:.2f}, Margem: {linha[6]:.2f}%\n")
        else:
            text_area.insert(tk.END, "Nenhum dado encontrado.\n")
    except FileNotFoundError:
        messagebox.showerror("Erro", "Arquivo 'planilha.xlsx' não encontrado!")

# Função para deletar produto
def deletar_produto():
    nome_produto = entry_deletar.get()
    if not nome_produto:
        messagebox.showerror("Erro", "Digite o nome do produto para deletar!")
        return
    
    try:
        deletar_linha_por_nome(nome_produto, 'planilha.xlsx')
        messagebox.showinfo("Sucesso", f"Produto '{nome_produto}' deletado!")
        carregar_dados()  # Atualizar exibição
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao deletar: {str(e)}")

# Funções auxiliares (baseadas no código original, corrigidas)
def obter_dados_excel(nome_planilha):
    wb = openpyxl.load_workbook(nome_planilha)
    sheet = wb.active
    dados = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        dados.append(row)
    return dados

def deletar_linha_por_nome(nome_produto, nome_planilha):
    wb = openpyxl.load_workbook(nome_planilha)
    sheet = wb.active
    contador = 2
    deletado = False
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        if str(row[0]).strip() == nome_produto.strip():
            sheet.delete_rows(contador)
            deletado = True
            break
        contador += 1
    if not deletado:
        raise ValueError(f"Produto '{nome_produto}' não encontrado.")
    wb.save(nome_planilha)

# Variável global para armazenar cálculos
valores_calculados = None

# Criar GUI
root = tk.Tk()
root.title("Calculadora de Lucro com Gráficos em Excel")
root.geometry("500x600")

# Campos de entrada
tk.Label(root, text="Nome do Produto:").pack(pady=5)
entry_nome = tk.Entry(root)
entry_nome.pack(pady=5)

tk.Label(root, text="Preço de Compra:").pack(pady=5)
entry_compra = tk.Entry(root)
entry_compra.pack(pady=5)

tk.Label(root, text="Preço de Venda:").pack(pady=5)
entry_venda = tk.Entry(root)
entry_venda.pack(pady=5)

tk.Label(root, text="Custos Adicionais (opcional):").pack(pady=5)
entry_custos = tk.Entry(root)
entry_custos.pack(pady=5)

tk.Label(root, text="Custo do Frete (opcional):").pack(pady=5)
entry_frete = tk.Entry(root)
entry_frete.pack(pady=5)

# Botão calcular
tk.Button(root, text="Calcular Lucro", command=calcular_lucro).pack(pady=10)
label_resultado = tk.Label(root, text="")
label_resultado.pack(pady=5)

# Botão salvar
tk.Button(root, text="Salvar na Planilha (com Gráfico)", command=salvar_dados).pack(pady=10)

# Área para exibir dados
tk.Button(root, text="Carregar e Exibir Dados", command=carregar_dados).pack(pady=10)
text_area = tk.Text(root, height=10, width=60)
text_area.pack(pady=10)

# Deletar produto
tk.Label(root, text="Nome do Produto para Deletar:").pack(pady=5)
entry_deletar = tk.Entry(root)
entry_deletar.pack(pady=5)
tk.Button(root, text="Deletar Produto", command=deletar_produto).pack(pady=10)

# Iniciar GUI
root.mainloop()
