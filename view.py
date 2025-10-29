# importando openxl
import openpyxl

# nome da folha excel
nome = 'planilha.xlsx'  # corrigido nome do arquivo

# obtendo os dados
def obter_dados_excel(nome_planilha):
    wb = openpyxl.load_workbook(nome_planilha)
    sheet = wb.active
    dados = []

    for row in sheet.iter_rows(min_row=2, values_only=True):  # corrigido 'is' para 'in'
        dados.append(row)
    
    return dados


#deletar produto
def deletar_line_por_nome(nome_produto, nome_planilha):
     wb = openpyxl.load_workbook(nome_planilha)
    sheet = wb.active
    contador = 2

    for row in sheet.inter_row(min_row=2,min_col=1, max_col=1, values_only=True):
        if str(row[0]) == nome_produto:
            #Obter o numero da linha e deletar a linha inteira
            linha = contador
            sheet.delete_rows(linha)
            break

            # incrementando o contador
            contador +=1

            # Salvando as mudanças na planilha
            wb.save(nome_planilha)



deletar_line_por_nome('produto 2',nome)