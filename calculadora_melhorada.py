import tkinter as tk
from tkinter import messagebox, ttk
import os
import subprocess

# Verificar dependências ao iniciar
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
except ImportError:
    messagebox.showerror("Erro de Dependência", "Biblioteca 'openpyxl' não instalada. Instale com: pip install openpyxl")
    exit()

try:
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
except ImportError:
    messagebox.showerror("Erro de Dependência", "Biblioteca 'matplotlib' não instalada. Instale com: pip install matplotlib")
    exit()

class CalculadoraLucroApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Calculadora de Lucro Avançada - Tela Interativa")
        self.root.geometry("800x700")
        self.root.configure(bg="#f0f0f0")
        
        self.dados_atuais = []
        self.historico = []
        
        # Frames
        self.frame_entrada = tk.Frame(root, bg="#f0f0f0", padx=10, pady=10)
        self.frame_entrada.pack(fill="x")
        
        self.frame_resultados = tk.Frame(root, bg="#f0f0f0", padx=10, pady=10)
        self.frame_resultados.pack(fill="x")
        
        self.frame_acoes = tk.Frame(root, bg="#f0f0f0", padx=10, pady=10)
        self.frame_acoes.pack(fill="x")
        
        self.frame_tabela = tk.Frame(root, bg="#f0f0f0", padx=10, pady=10)
        self.frame_tabela.pack(fill="both", expand=True)
        
        self.criar_widgets()
        print("App inicializado com sucesso.")  # Debug
    
    def criar_widgets(self):
        try:
            # Entrada
            tk.Label(self.frame_entrada, text="📝 Entrada de Dados", font=("Arial", 14, "bold"), bg="#f0f0f0").grid(row=0, column=0, columnspan=2, pady=5)
            
            tk.Label(self.frame_entrada, text="Nome do Produto:", bg="#f0f0f0").grid(row=1, column=0, sticky="w")
            self.entry_nome = tk.Entry(self.frame_entrada, width=30)
            self.entry_nome.grid(row=1, column=1, pady=2)
            
            tk.Label(self.frame_entrada, text="Preço de Compra (R$):", bg="#f0f0f0").grid(row=2, column=0, sticky="w")
            self.entry_compra = tk.Entry(self.frame_entrada, width=30)
            self.entry_compra.grid(row=2, column=1, pady=2)
            
            tk.Label(self.frame_entrada, text="Preço de Venda (R$):", bg="#f0f0f0").grid(row=3, column=0, sticky="w")
            self.entry_venda = tk.Entry(self.frame_entrada, width=30)
            self.entry_venda.grid(row=3, column=1, pady=2)
            
            tk.Label(self.frame_entrada, text="Custos Adicionais (R$):", bg="#f0f0f0").grid(row=4, column=0, sticky="w")
            self.entry_custos = tk.Entry(self.frame_entrada, width=30)
            self.entry_custos.grid(row=4, column=1, pady=2)
            
            tk.Label(self.frame_entrada, text="Custo do Frete (R$):", bg="#f0f0f0").grid(row=5, column=0, sticky="w")
            self.entry_frete = tk.Entry(self.frame_entrada, width=30)
            self.entry_frete.grid(row=5, column=1, pady=2)
            
            tk.Button(self.frame_entrada, text="➕ Adicionar Produto", command=self.adicionar_produto, bg="#4CAF50", fg="white").grid(row=6, column=0, pady=10)
            tk.Button(self.frame_entrada, text="🔄 Resetar Campos", command=self.resetar_campos, bg="#FF9800", fg="white").grid(row=6, column=1, pady=10)
            
            # Resultados
            tk.Label(self.frame_resultados, text="📊 Resultados e Histórico", font=("Arial", 14, "bold"), bg="#f0f0f0").pack()
            self.label_result = tk.Label(self.frame_resultados, text="", bg="#f0f0f0", font=("Arial", 10))
            self.label_result.pack()
            
            self.listbox_historico = tk.Listbox(self.frame_resultados, height=5, width=80)
            self.listbox_historico.pack(pady=5)
            
            # Gráfico
            self.fig, self.ax = plt.subplots(figsize=(5, 3))
            self.canvas = FigureCanvasTkAgg(self.fig, master=self.frame_resultados)
            self.canvas.get_tk_widget().pack()
            
            # Ações
            tk.Label(self.frame_acoes, text="⚙️ Ações", font=("Arial", 14, "bold"), bg="#f0f0f0").pack()
            tk.Button(self.frame_acoes, text="💾 Salvar em Excel", command=self.salvar_excel, bg="#2196F3", fg="white").pack(side="left", padx=5)
            tk.Button(self.frame_acoes, text="📂 Carregar de Excel", command=self.carregar_excel, bg="#2196F3", fg="white").pack(side="left", padx=5)
            tk.Button(self.frame_acoes, text="🗑️ Deletar Produto", command=self.deletar_produto, bg="#F44336", fg="white").pack(side="left", padx=5)
            tk.Button(self.frame_acoes, text="📈 Exportar Gráfico", command=self.exportar_grafico, bg="#9C27B0", fg="white").pack(side="left", padx=5)
            tk.Button(self.frame_acoes, text="📖 Abrir Excel", command=self.abrir_excel, bg="#607D8B", fg="white").pack(side="left", padx=5)
            
            # Tabela
            self.tree = ttk.Treeview(self.frame_tabela, columns=("Nome", "Compra", "Venda", "Custos", "Frete", "Lucro", "Margem"), show="headings", height=10)
            self.tree.pack(fill="both", expand=True)
            for col in self.tree["columns"]:
                self.tree.heading(col, text=col)
                self.tree.column(col, width=100)
            
            scrollbar = ttk.Scrollbar(self.frame_tabela, orient="vertical", command=self.tree.yview)
            self.tree.configure(yscrollcommand=scrollbar.set)
            scrollbar.pack(side="right", fill="y")
            
            print("Widgets criados com sucesso.")  # Debug
        except Exception as e:
            messagebox.showerror("Erro ao Criar Interface", f"Erro: {str(e)}")
    
    def adicionar_produto(self):
        try:
            nome = self.entry_nome.get().strip()
            if not nome:
                raise ValueError("Nome do produto é obrigatório.")
            
            preco_compra = float(self.entry_compra.get())
            preco_venda = float(self.entry_venda.get())
            custos = float(self.entry_custos.get() or 0)
            frete = float(self.entry_frete.get() or 0)
            
            if preco_compra < 0 or preco_venda < 0 or custos < 0 or frete < 0:
                raise ValueError("Valores não podem ser negativos.")
            if preco_venda == 0:
                raise ValueError("Preço de venda não pode ser zero.")
            
            lucro = preco_venda - preco_compra - custos - frete
            margem = (lucro / preco_venda) * 100
            
            produto = [nome, preco_compra, preco_venda, custos, frete, lucro, margem]
            self.dados_atuais.append(produto)
            
            self.historico.append(f"{nome}: Lucro R$ {lucro:.2f}, Margem {margem:.2f}%")
            self.listbox_historico.insert(tk.END, self.historico[-1])
            
            self.label_result.config(text=f"Produto '{nome}' adicionado!")
            self.resetar_campos()
            print(f"Produto '{nome}' adicionado.")  # Debug
        except ValueError as e:
            messagebox.showerror("Erro de Validação", str(e))
        except Exception as e:
            messagebox.showerror("Erro Inesperado", f"Erro: {str(e)}")
    
    def resetar_campos(self):
        try:
            self.entry_nome.delete(0, tk.END)
            self.entry_compra.delete(0, tk.END)
            self.entry_venda.delete(0, tk.END)
            self.entry_custos.delete(0, tk.END)
            self.entry_frete.delete(0, tk.END)
        except Exception as e:
            messagebox.showerror("Erro ao Resetar", f"Erro: {str(e)}")
    
    def salvar_excel(self):
        if not self.dados_atuais:
            messagebox.showerror("Erro", "Adicione produtos antes de salvar!")
            return
        
        try:
            try:
                wb = openpyxl.load_workbook('planilha.xlsx')
                sheet = wb.active
            except FileNotFoundError:
                wb = Workbook()
                sheet = wb.active
                sheet.title = 'Calculadora de Lucro'
                sheet['A1'] = 'Nome do Produto'
                sheet['B1'] = 'Preço de Compra'
                sheet['C1'] = 'Preço de Venda'
                sheet['D1'] = 'Custos Adicionais'
                sheet['E1'] = 'Custo do Frete'
                sheet['F1'] = 'Lucro Líquido'
                sheet['G1'] = 'Margem de Lucro (%)'
            
            for produto in self.dados_atuais:
                row = sheet.max_row + 1
                for i, val in enumerate(produto):
                    sheet.cell(row=row, column=i+1, value=val)
            
            # Gráfico
            chart = BarChart()
            chart.title = "Lucros dos Produtos"
            data = Reference(sheet, min_col=6, min_row=2, max_row=sheet.max_row)
            cats = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
            chart.add_data(data)
            chart.set_categories(cats)
            sheet.add_chart(chart, "H2")
            
            wb.save('planilha.xlsx')
            self.dados_atuais.clear()
            messagebox.showinfo("Sucesso", "Salvo em 'planilha.xlsx'!")
            print("Dados salvos.")  # Debug
        except Exception as e:
            messagebox.showerror("Erro ao Salvar", f"Erro: {str(e)}")
    
    def carregar_excel(self):
        try:
            wb = openpyxl.load_workbook('planilha.xlsx')
            sheet = wb.active
            self.tree.delete(*self.tree.get_children())
            nomes = []
            lucros = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) >= 7:
                    self.tree.insert("", tk.END, values=row)
                    nomes.append(str(row[0]) if row[0] else "Sem Nome")
                    lucros.append(float(row[5]) if row[5] is not None else 0.0)
            
            if nomes and lucros:
                self.ax.clear()
                self.ax.bar(nomes, lucros, color='skyblue')
                self.ax.set_title("Lucros dos Produtos")
                self.ax.set_xlabel("Produtos")
                self.ax.set_ylabel("Lucro (R$)")
                self.canvas.draw()
            else:
                self.ax.clear()
                self.ax.text(0.5, 0.5, "Nenhum dado", ha='center', va='center', transform=self.ax.transAxes)
                self.canvas.draw()
            print("Dados carregados.")  # Debug
        except FileNotFoundError:
            messagebox.showerror("Erro", "Arquivo 'planilha.xlsx' não encontrado!")
        except Exception as e:
            messagebox.showerror("Erro ao Carregar", f"Erro: {str(e)}")
    
    def deletar_produto(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showerror("Erro", "Selecione um produto!")
            return
        
        try:
            nome = self.tree.item(selected[0], "values")[0]
            wb = openpyxl.load_workbook('planilha.xlsx')
            sheet = wb.active
            contador = 2
            deletado = False
            while contador <= sheet.max_row:
                cell_value = sheet.cell(row=contador, column=1).value
                if cell_value and str(cell_value).strip() == nome.strip():
                    sheet.delete_rows(contador)
                    deletado = True
                    break
                contador += 1
            
            if deletado:
                wb.save('planilha.xlsx')
                messagebox.showinfo("Sucesso", f"Produto '{nome}' deletado!")
                self.carregar_excel()
                print(f"Produto '{nome}' deletado.")  # Debug
            else:
                messagebox.showerror("Erro", "Produto não encontrado!")
        except FileNotFoundError:
            messagebox.showerror("Erro", "Arquivo não encontrado!")
        except Exception as e:
            messagebox.showerror("Erro ao Deletar", f"Erro: {str(e)}")
    
    def exportar_grafico(self):
        try:
            if not self.ax.get_children():
                raise ValueError("Nenhum gráfico para exportar.")
            self.fig.savefig('grafico_lucros.png')
            messagebox.showinfo("Sucesso", "Gráfico salvo como 'grafico_lucros.png'!")
            print("Gráfico exportado.")  # Debug
        except Exception as e:
            messagebox.showerror("Erro ao Exportar", f"Erro: {str(e)}")
    
    def abrir_excel(self):
        try:
            if os.name == 'nt':
                os.startfile('planilha.xlsx')
            elif os.name == 'posix':
                subprocess.run(['xdg-open', 'planilha.xlsx'])  # Linux
            else:
                subprocess.run(['open', 'planilha.xlsx'])  # macOS
            print("Excel aberto.")  # Debug
        except FileNotFoundError:
            messagebox.showerror("Erro", "Arquivo não encontrado!")
        except Exception as e:
            messagebox.showerror("Erro ao Abrir", f"Erro: {str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = CalculadoraLucroApp(root)
    root.mainloop()
