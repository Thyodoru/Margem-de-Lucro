import tkinter as tk
from tkinter import messagebox, filedialog
import openpyxl
from openpyxl import Workbook

# Função para salvar dados em Excel
def salvar_dados():
    nome = entry_nome.get()
    idade = entry_idade.get()
    cidade = entry_cidade.get()
    
    if not nome or not idade or not cidade:
        messagebox.showerror("Erro", "Preencha todos os campos!")
        return
    
    # Criar um novo workbook e planilha
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados Pessoais"
    
    # Adicionar cabeçalhos
    ws['A1'] = "Nome"
    ws['B1'] = "Idade"
    ws['C1'] = "Cidade"
    
    # Adicionar dados
    ws['A2'] = nome
    ws['B2'] = idade
    ws['C2'] = cidade
    
    # Salvar o arquivo
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        wb.save(file_path)
        messagebox.showinfo("Sucesso", f"Dados salvos em {file_path}")

# Função para carregar dados de Excel
def carregar_dados():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # Ler dados (assumindo estrutura simples)
            dados = []
            for row in ws.iter_rows(min_row=2, values_only=True):  # Pula cabeçalho
                dados.append(row)
            
            # Exibir na tela (limpa e adiciona)
            text_area.delete(1.0, tk.END)
            if dados:
                for linha in dados:
                    text_area.insert(tk.END, f"Nome: {linha[0]}, Idade: {linha[1]}, Cidade: {linha[2]}\n")
            else:
                text_area.insert(tk.END, "Nenhum dado encontrado na planilha.\n")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar arquivo: {str(e)}")

# Criar a janela principal
root = tk.Tk()
root.title("Tela Interativa com Openpyxl")
root.geometry("400x400")

# Labels e entradas
tk.Label(root, text="Nome:").pack(pady=5)
entry_nome = tk.Entry(root)
entry_nome.pack(pady=5)

tk.Label(root, text="Idade:").pack(pady=5)
entry_idade = tk.Entry(root)
entry_idade.pack(pady=5)

tk.Label(root, text="Cidade:").pack(pady=5)
entry_cidade = tk.Entry(root)
entry_cidade.pack(pady=5)

# Botões
tk.Button(root, text="Salvar em Excel", command=salvar_dados).pack(pady=10)
tk.Button(root, text="Carregar de Excel", command=carregar_dados).pack(pady=10)

# Área de texto para exibir dados carregados
text_area = tk.Text(root, height=10, width=40)
text_area.pack(pady=10)

# Iniciar o loop da interface
root.mainloop()
