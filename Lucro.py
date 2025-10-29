# importando openxl
import openpyxl


# definindo as variaveis
preco_compra = float(input('Digite o preço de compra: '))
preco_venda = float(input('Digite o preço de venda: '))
custos_adicionais = float(input('Digite os custos adicionais (opcional):  '))
custo_frete = float(input('Digite o custo do frete (opcional): '))

# Calculando o lucro liquido
lucro = preco_venda - preco_compra - custos_adicionais

#imrpimindo o resultado
print(f'O lucro líquido da venda é: R$ {lucro:.2f}')

# calculando margem de lucro
margem_lucro = (lucro / preco_venda) * 100

#imprimindo o resultado
print('A margem de lucro é de {:.2f}%'.format(margem_lucro))


# Salvando resultados em uma planilha Excel

nome_do_produto = 'produto 1'

resumo = [nome_do_produto, preco_compra, preco_venda, custos_adicionais, custo_frete, lucro, margem_lucro]

# carregando a planilha existente ou criando uma nova planilha se ela não existe
try:
    wb =openpyxl.load_workbook('planilha.xlsx')
    sheet = wb.active
except:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'resumo da calculadora de lucro' 

    #adicionando cabeçalho
    sheet['A1'] = 'Nome do Produto'
    sheet['B1'] = 'Preço de Compra'
    sheet['C1'] = 'Preço de Venda'
    sheet['D1'] = 'Custos Adicionais'
    sheet['E1'] = 'Custo do Frete'
    sheet['F1'] = 'Lucro Líquido'
    sheet['G1'] = 'Margem de Lucro (%)'
  
#Adicionando valors a planilha
row = sheet.max_row + 1
sheet['A{}'.format(row)] = nome_do_produto
sheet['B{}'.format(row)] = preco_compra
sheet['C{}'.format(row)] = preco_venda
sheet['D{}'.format(row)] = custos_adicionais
sheet['E{}'.format(row)] = custo_frete
sheet['F{}'.format(row)] = lucro
sheet['G{}'.format(row)] = margem_lucro

#salvando a planilha
wb.save('planilha.xlsx')

